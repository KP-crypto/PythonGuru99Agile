import openpyxl

def getrowcount(file,sheetname):
    wb=openpyxl.load_workbook(file)
    sheet=wb[sheetname]
    rows=sheet.max_row
    return rows

def getcolumncount(file,sheetname):
    wb = openpyxl.load_workbook(file)
    sheet = wb[sheetname]
    columns = sheet.max_column
    return columns

def readexel(file,sheetname,rownumber,colnumber):
    wb = openpyxl.load_workbook(file)
    sheet = wb[sheetname]
    return sheet.cell(row=rownumber,column=colnumber).value

def writeexel(file,sheetname,rownumner,colnumber,data):
    wb = openpyxl.load_workbook(file)
    sheet = wb[sheetname]
    sheet.cell(row=rownumner,column=colnumber).value=data
    wb.save(file)





